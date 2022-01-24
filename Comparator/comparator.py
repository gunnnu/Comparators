from distutils.log import error
import os
from platform import system
import sys
import openpyxl as xl
from openpyxl.styles import PatternFill, Font, Alignment
import json
import pandas as pd

def widthAdjust(wksheet):
	ws = wksheet
	dims = {}
	for row in ws.rows:
		for cell in row:
			if cell.value:
				dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
	for col, value in dims.items():
		ws.column_dimensions[col].width = value+4



def openFile(filename):
	if(sys.platform=="linux"):
		os.system(f"xdg-open {filename}")
	elif(sys.platform=="darwin"):
		os.system(f"open {filename}")
	else:
		os.system(f'start "" "{filename}"')


def colorit(string, status):
    attr = []
    if status:
        # green
        attr.append('1;42')
    else:
        # red
        attr.append('1;41')
    return '\x1b[%sm%s\x1b[0m' % (';'.join(attr), f" {string} ")

ErrorStr=colorit("Error :", 0)
print()

args = sys.argv
# print(len(args))

if (len(args) > 2):
    path1 = args[1]
    path2 = args[2]
elif (len(args) < 2):
    print(f"{ErrorStr} No file specified")
    exit(1)
else:
    print(f"{ErrorStr} Please provide second file too")
    exit(1)
try:
	wb = xl.Workbook()
	sheet_a = wb.active
	sheet_a.title="Source Data"
	sheet_b = wb.create_sheet(title="File to be Compared")

	wb1 = xl.load_workbook(path1)
	sheet1 = wb1.active
	
	wb2 = xl.load_workbook(path2)
	sheet2 = wb2.active

    # for i in range(1,3):
	jsonData = pd.read_excel(locals()[f"path{1}"])
    # print(sheet2.max_column)
    # for row1 in sheet2:
    # 	for row2 in sheet1:
	for j in range(1, sheet1.max_row+1):
		for i in range(1, sheet1.max_column+1):
			if(sheet1[j][i-1].value == sheet2[j][i-1].value):
				print(colorit("Match", 1), end=" | ")
				res="pass"
			else:
				print(colorit("Do Not match", 0), end=" | ")                # print(f": {sheet1[j][i].value} = {sheet2[j][i].value} X", end=" | ")
				res="fail"
			sheet_a.cell(column=i, row=j, value=f"{sheet1[j][i-1].value}")
			sheet_b.cell(column=i, row=j, value=f"{sheet2[j][i-1].value}")
			if (res=="fail"):
				sheet_a[j][i-1].fill = PatternFill("solid", start_color="FF0000")
				sheet_b[j][i-1].fill = PatternFill("solid", start_color="FF0000")

			sheet_a[j][i-1].alignment = Alignment(horizontal="center", vertical="center")
			sheet_b[j][i-1].alignment = Alignment(horizontal="center", vertical="center")
		print()

        # locals()[f"jsonData{i}"]=jsonData.to_json(orient='records',date_format='epoch' ,indent=4).replace('\\','')
        # f=open("op.json","w")
        # f.write(jsonData)
    # print(locals()["jsonData2"])
	# print(sys.platform)
	print("\n")
	for i in range(1,sheet_a.max_column+1):
		sheet_a[1][i-1].font=Font(bold=True)
		sheet_b[1][i-1].font=Font(bold=True)
		
	widthAdjust(sheet_a)
	widthAdjust(sheet_b)
	try:
		wb.save(filename = "Results.xlsx")
		openFile("Results.xlsx")
	except PermissionError as err:
		print(f"{ErrorStr} Permission denied to open the file. The Results file is Already open, please close it first.")
		print(f"Error details :\n	=>{err}\n")
except FileNotFoundError as err:
    print(f"{ErrorStr} One of the files was not found :")
    print(f"	=>{err}\n")
except:
    print(f"{ErrorStr} Following Error Occured :")
    print(f"	=>{err}\n")
